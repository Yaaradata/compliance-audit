"""Parse GCP_Evidence_CollectionforSWIFT_v2026_Updated.xlsx — sheet 1 evidence mapping."""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

# Sheet 1 header row (1-based) — "SWIFT Domain" in column A
_SHEET_NAME = "1_Evidence_GCP_Mapping"
_HEADER_MARKERS = ("SWIFT Domain", "Evidence Item", "GCP Service")


@dataclass(frozen=True)
class WorkbookEvidenceRow:
    swift_domain: str
    evidence_item_raw: str
    item_code: str | None
    gcp_asset_resource: str | None
    security_configuration: str | None
    vulnerability_signal: str | None
    gcp_service: str | None
    api_data_source: str | None
    automation_feasibility: str | None


def _norm_cell(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _parse_item_code(evidence_cell: str | None) -> str | None:
    """Extract A1, A2, B3, A10 from strings like 'A1 – Network' or 'A1 - Network'."""
    if not evidence_cell:
        return None
    m = re.match(r"^\s*([A-Z]\d+[A-Z]?)\b", evidence_cell.strip(), re.IGNORECASE)
    if not m:
        return None
    return m.group(1).strip().upper()


def load_workbook_evidence_mapping(workbook_path: str | Path) -> list[WorkbookEvidenceRow]:
    path = Path(workbook_path)
    if not path.is_file():
        raise FileNotFoundError(f"GCP workbook not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if _SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Workbook missing sheet {_SHEET_NAME!r}; found {wb.sheetnames!r}")

    ws = wb[_SHEET_NAME]
    header_row_idx: int | None = None
    rows_iter = ws.iter_rows(values_only=True)
    for idx, row in enumerate(rows_iter, start=1):
        cells = [(_norm_cell(c) or "") for c in (row or ())]
        if len(cells) >= 2 and cells[0] == _HEADER_MARKERS[0] and _HEADER_MARKERS[1] in cells[1]:
            header_row_idx = idx
            break
        if len(cells) >= 2 and cells[0].startswith(_HEADER_MARKERS[0]) and "Evidence" in cells[1]:
            header_row_idx = idx
            break

    if header_row_idx is None:
        wb.close()
        raise ValueError(f"Could not locate header row in {_SHEET_NAME}")

    # Re-read from header — read_only sheet may not support seek; reload
    wb.close()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[_SHEET_NAME]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row = all_rows[header_row_idx - 1]
    header = [(_norm_cell(c) or "").lower() for c in header_row]

    def col_index(*needles: str) -> int | None:
        for i, h in enumerate(header):
            for n in needles:
                if n.lower() in h:
                    return i
        return None

    ix_domain = col_index("swift domain", "domain")
    ix_item = col_index("evidence item")
    ix_asset = col_index("gcp asset", "resource")
    ix_sec = col_index("security configuration", "security")
    ix_vuln = col_index("vulnerability")
    ix_svc = col_index("gcp service", "service")
    ix_api = col_index("api", "data source")
    ix_auto = col_index("automation")

    if ix_item is None:
        raise ValueError("Workbook: missing Evidence Item column")

    out: list[WorkbookEvidenceRow] = []
    last_domain = ""
    for row in all_rows[header_row_idx:]:
        if not row:
            continue
        cells = list(row)
        while len(cells) < len(header_row):
            cells.append(None)

        def get(i: int | None) -> str | None:
            if i is None or i >= len(cells):
                return None
            return _norm_cell(cells[i])

        item_raw = get(ix_item)
        if not item_raw:
            continue

        d = get(ix_domain) or ""
        if d:
            last_domain = d
        domain = last_domain
        code = _parse_item_code(item_raw)

        out.append(
            WorkbookEvidenceRow(
                swift_domain=domain,
                evidence_item_raw=item_raw,
                item_code=code,
                gcp_asset_resource=get(ix_asset),
                security_configuration=get(ix_sec),
                vulnerability_signal=get(ix_vuln),
                gcp_service=get(ix_svc),
                api_data_source=get(ix_api),
                automation_feasibility=get(ix_auto),
            )
        )

    return out


def resolve_workbook_path(explicit: str | Path | None = None) -> Path | None:
    """Path to the Updated workbook: explicit arg, env GCP_EVIDENCE_WORKBOOK_PATH, or repo root default."""
    if explicit:
        p = Path(explicit)
        return p if p.is_file() else None
    try:
        from app.config import settings

        p = (getattr(settings, "GCP_EVIDENCE_WORKBOOK_PATH", None) or "").strip()
        if p:
            path = Path(p)
            if path.is_file():
                return path
    except Exception:
        pass
    candidate = Path(__file__).resolve().parents[4] / "GCP_Evidence_CollectionforSWIFT_v2026_Updated.xlsx"
    return candidate if candidate.is_file() else None
