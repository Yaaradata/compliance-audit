"""
Load CSCF_v2025_Complete_Sufficiency_Matrix.xlsx into evidence_sufficiency_matrix.

Sheet columns (any case/newlines normalized):
  Item Code
  Evidence Item Name
  Control ID
  Control Name
  M/A
  Evidence Type
  What This Evidence Must Show for This Control (Sufficiency Criteria — Bullet Points)
  How to Evaluate / AI Scoring Criteria (Pass/Fail Checkpoints)

DB columns:
  item_code, control_id, evidence_item_name, control_name, ma, evidence_type,
  sufficiency_criteria, evaluation_criteria, cscf_version, created_at

Usage:
  cd backend
  pip install openpyxl psycopg2-binary python-dotenv
  python scripts/load_evidence_sufficiency_matrix.py

Optional env:
  XLSX_FILE  — path to xlsx (default: project_root/CSCF_v2025_Complete_Sufficiency_Matrix.xlsx)
  DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD, DB_SSL
"""

import json
import os
import re
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
PROJECT_ROOT = BACKEND_DIR.parent

# Load .env from backend
try:
    from dotenv import load_dotenv
    load_dotenv(BACKEND_DIR / ".env")
except ImportError:
    pass

DEFAULT_XLSX = PROJECT_ROOT / "CSCF_v2025_Complete_Sufficiency_Matrix.xlsx"
SCHEMA = "swift_2025"
CSCF_VERSION = "2025v"

# Sheet header → DB field. Order matters: first match wins.
# Headers are normalized (lowercase, newlines → space) before matching.
SHEET_TO_DB = [
    ("item code", "item_code"),
    ("evidence item name", "evidence_item_name"),
    ("control id", "control_id"),
    ("control name", "control_name"),
    ("m/a", "ma"),
    ("evidence type", "evidence_type"),
    ("what this evidence must show for this control", "sufficiency_criteria"),
    ("sufficiency criteria", "sufficiency_criteria"),
    ("how to evaluate", "evaluation_criteria"),
    ("ai scoring criteria", "evaluation_criteria"),
    ("pass/fail checkpoints", "evaluation_criteria"),
]

# DB max lengths (evidence_sufficiency_matrix)
ITEM_CODE_MAX = 5
CONTROL_ID_MAX = 10
EVIDENCE_ITEM_NAME_MAX = 255
CONTROL_NAME_MAX = 255
MA_MAX = 1
EVIDENCE_TYPE_MAX = 100


def normalize_header(h: str) -> str:
    """Lowercase, replace newlines/tabs with space, collapse spaces, remove punctuation for matching."""
    if h is None:
        return ""
    s = str(h).strip().replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    # Remove punctuation so "Item Code" and "Item Code:" match
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def safe_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def bullet_text_to_json(text: str | None) -> str | None:
    """
    Convert bullet-point text to a JSON string with keys "1", "2", "3", ...
    Splits by newlines and by bullet char (•); each resulting line is one point.
    Leading bullets (•, -, *, ·, or "N. ") are stripped.
    Returns None if text is empty; otherwise json.dumps({"1": "...", "2": "...", ...}).
    """
    if not text or not str(text).strip():
        return None
    raw = str(text).strip()
    points = []
    for line in re.split(r"[\r\n]+", raw):
        # Split by bullet so "• A • B" on one line gives two points
        for part in re.split(r"\s*[•]\s*", line):
            part = part.strip()
            if not part:
                continue
            part = re.sub(r"^[\s\-*·]+\s*", "", part)
            part = re.sub(r"^\d+[.)]\s*", "", part)
            part = part.strip()
            if part:
                points.append(part)
    if not points:
        return None
    obj = {str(i): p for i, p in enumerate(points, start=1)}
    return json.dumps(obj)


def get_connection():
    import psycopg2
    return psycopg2.connect(
        host=os.getenv("DB_HOST", "127.0.0.1"),
        port=int(os.getenv("DB_PORT", "5432")),
        dbname=os.getenv("DB_NAME", "compliance"),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASSWORD", ""),
        sslmode="require" if os.getenv("DB_SSL", "false").lower() == "true" else "disable",
    )


def build_header_map(headers: list[str]) -> dict[str, str]:
    """
    Map each DB field to the first sheet header that matches.
    Returns dict: db_field -> original header string (key for row_dict).
    """
    # Normalize and dedupe: norm_key -> first original header that normalizes to it
    norm_to_orig: dict[str, str] = {}
    for h in headers:
        if h is None:
            continue
        orig = str(h).strip() if h else ""
        if not orig:
            continue
        key = normalize_header(orig)
        if key and key not in norm_to_orig:
            norm_to_orig[key] = orig

    # Match: for each (sheet_substring, db_field) find first norm that contains or is contained in substring
    sheet_substrings_norm = [(normalize_header(ss), db_field) for ss, db_field in SHEET_TO_DB]
    db_to_sheet: dict[str, str] = {}
    for sub_norm, db_field in sheet_substrings_norm:
        if db_field in db_to_sheet:
            continue
        for norm, orig in norm_to_orig.items():
            if not norm:
                continue
            if sub_norm in norm or norm in sub_norm:
                db_to_sheet[db_field] = orig
                break
    return db_to_sheet


def parse_xlsx(filepath: Path) -> list[dict]:
    """
    Read first sheet of xlsx and return list of dicts with keys:
    item_code, control_id, evidence_item_name, control_name, ma, evidence_type,
    sufficiency_criteria, evaluation_criteria.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row: first row with at least 2 non-empty cells; prefer row whose normalized cells match known columns
    header_row_idx = 1
    raw_headers = None
    for row_num in range(1, min(6, ws.max_row + 1)):
        row_cells = [c.value for c in next(ws.iter_rows(min_row=row_num, max_row=row_num))]
        row_strs = [safe_str(v) or "" for v in row_cells]
        non_empty = [s for s in row_strs if s]
        if len(non_empty) >= 2:
            raw_headers = row_cells
            header_row_idx = row_num
            break
    if not raw_headers:
        wb.close()
        raise ValueError("Sheet has no row with at least 2 non-empty cells in the first 5 rows.")

    headers = [safe_str(h) if h is not None else "" for h in raw_headers]
    row_dict_keys = [h if h else f"_col{i}" for i, h in enumerate(headers)]  # use placeholder for empty headers
    db_to_sheet = build_header_map([h for h in headers if h])

    required = {"item_code", "control_id"}
    if not required.issubset(db_to_sheet):
        wb.close()
        first_row_preview = [str(v)[:40] for v in (raw_headers or [])[:8]]
        raise ValueError(
            f"Sheet must have columns matching 'Item Code' and 'Control ID'. "
            f"Found header map: {db_to_sheet}. "
            f"First row (row {header_row_idx}) preview: {first_row_preview}"
        )

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        row_dict = dict(zip(row_dict_keys, row))
        item_code = safe_str(row_dict.get(db_to_sheet["item_code"]))
        control_id = safe_str(row_dict.get(db_to_sheet["control_id"]))
        if not item_code or not control_id:
            continue

        evidence_item_name = (safe_str(row_dict.get(db_to_sheet.get("evidence_item_name"))) or "")[:EVIDENCE_ITEM_NAME_MAX]
        control_name = (safe_str(row_dict.get(db_to_sheet.get("control_name"))) or "")[:CONTROL_NAME_MAX]
        ma_raw = safe_str(row_dict.get(db_to_sheet.get("ma"))) or "M"
        ma = (ma_raw[:1].upper() if ma_raw else "M")
        if ma not in ("M", "A"):
            ma = "M"
        evidence_type = (safe_str(row_dict.get(db_to_sheet.get("evidence_type"))) or "")[:EVIDENCE_TYPE_MAX]
        sufficiency_raw = safe_str(row_dict.get(db_to_sheet.get("sufficiency_criteria")))
        evaluation_raw = safe_str(row_dict.get(db_to_sheet.get("evaluation_criteria")))
        # Store as JSON string: {"1": "point one", "2": "point two", ...}
        sufficiency_criteria = bullet_text_to_json(sufficiency_raw)
        evaluation_criteria = bullet_text_to_json(evaluation_raw)

        rows.append({
            "item_code": (item_code or "")[:ITEM_CODE_MAX],
            "control_id": (control_id or "")[:CONTROL_ID_MAX],
            "evidence_item_name": evidence_item_name,
            "control_name": control_name,
            "ma": ma[:MA_MAX],
            "evidence_type": evidence_type,
            "sufficiency_criteria": sufficiency_criteria,
            "evaluation_criteria": evaluation_criteria,
        })
    wb.close()
    return rows


def load_into_db(conn, rows: list[dict]) -> int:
    """Upsert all parsed rows into evidence_sufficiency_matrix (no FK checks; table has no references)."""
    if not rows:
        return 0
    cur = conn.cursor()
    sql = f"""
    INSERT INTO {SCHEMA}.evidence_sufficiency_matrix
        (item_code, control_id, evidence_item_name, control_name, ma, evidence_type,
         sufficiency_criteria, evaluation_criteria, cscf_version)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    ON CONFLICT (item_code, control_id) DO UPDATE SET
        evidence_item_name = EXCLUDED.evidence_item_name,
        control_name = EXCLUDED.control_name,
        ma = EXCLUDED.ma,
        evidence_type = EXCLUDED.evidence_type,
        sufficiency_criteria = EXCLUDED.sufficiency_criteria,
        evaluation_criteria = EXCLUDED.evaluation_criteria
    """
    count = 0
    for r in rows:
        # sufficiency_criteria and evaluation_criteria are already JSON strings from bullet_text_to_json
        cur.execute(sql, (
            r["item_code"],
            r["control_id"],
            r["evidence_item_name"],
            r["control_name"],
            r["ma"],
            r["evidence_type"],
            r.get("sufficiency_criteria"),   # JSON string or None
            r.get("evaluation_criteria"),    # JSON string or None
            CSCF_VERSION,
        ))
        count += 1
    conn.commit()
    cur.close()
    return count


def main():
    xlsx_path = Path(os.getenv("XLSX_FILE", str(DEFAULT_XLSX)))
    if not xlsx_path.is_absolute():
        xlsx_path = PROJECT_ROOT / xlsx_path
    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {xlsx_path.name}...")
    rows = parse_xlsx(xlsx_path)
    print(f"  Parsed {len(rows)} rows")

    if not rows:
        print("No rows to load. Exiting.")
        return

    print("Connecting to database...")
    conn = get_connection()
    try:
        n = load_into_db(conn, rows)
        print(f"Upserted {n} rows into {SCHEMA}.evidence_sufficiency_matrix")
    finally:
        conn.close()
    print("Done.")


if __name__ == "__main__":
    main()
