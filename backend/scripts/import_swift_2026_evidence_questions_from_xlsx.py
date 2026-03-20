#!/usr/bin/env python3
"""
Import SWIFT CSCF 2026 evidence questions from Excel/CSV into:
  swift_2026.evidence_based_questions

Expected input columns:
  - Control ID
  - Evidence Item
  - Question Key
  - Question (Label)
  - Question Type
  - Evidence Required
  - Evidence Source
  - Collection Method
  - AWS Auto Level
  - AWS Services
  - Question-Level AWS Sources
  - Reason / Rationale

Usage:
  python backend/scripts/import_swift_2026_evidence_questions_from_xlsx.py
  python backend/scripts/import_swift_2026_evidence_questions_from_xlsx.py --truncate-first
  python backend/scripts/import_swift_2026_evidence_questions_from_xlsx.py --file "ref-docs/swift/2026/SWIFT_CSCF2026_Evidence_Classification_v2.xlsx"
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook
from psycopg2.extras import execute_batch


SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
REPO_ROOT = BACKEND_DIR.parent

DEFAULT_INPUT_FILE = REPO_ROOT / "ref-docs" / "swift" / "2026" / "SWIFT_CSCF2026_Evidence_Classification_v2.xlsx"

EXPECTED_HEADERS = [
    "Control ID",
    "Evidence Item",
    "Question Key",
    "Question (Label)",
    "Question Type",
    "Evidence Required",
    "Evidence Source",
    "Collection Method",
    "AWS Auto Level",
    "AWS Services",
    "Question-Level AWS Sources",
    "Reason / Rationale",
]


def normalize_header(text: str) -> str:
    return " ".join((text or "").strip().lower().split())


HEADER_MAP = {normalize_header(h): h for h in EXPECTED_HEADERS}


def parse_bool(value: Any) -> bool:
    text = str(value or "").strip().lower()
    if text in {"1", "true", "yes", "y", "required", "mandatory"}:
        return True
    if text in {"0", "false", "no", "n", "optional", "not required"}:
        return False
    # Default to required=True for safety if the source value is ambiguous.
    return True


def parse_required(value: Any, question_type: str) -> bool:
    text = str(value or "").strip().lower()
    if text in {"1", "true", "yes", "y", "required", "mandatory"}:
        return True
    if text in {"0", "false", "no", "n", "optional", "not required"}:
        return False
    # In this workbook, "Evidence Required" often contains descriptors
    # like "Document/File Upload" rather than bool values.
    if question_type == "file":
        return False
    return True


def normalize_question_type(value: Any) -> str:
    raw = str(value or "").strip().lower().replace("_", " ").replace("-", " ")
    if raw in {"text", "short text", "input"}:
        return "text"
    if raw in {"textarea", "long text", "paragraph"}:
        return "textarea"
    if raw in {"file", "upload", "attachment"}:
        return "file"
    if raw in {"select", "dropdown", "single select", "single choice"}:
        return "select"
    if raw in {"multiselect", "multi select", "multiple choice"}:
        return "multiselect"
    if raw in {"date"}:
        return "date"
    if raw in {"checkbox", "boolean", "bool"}:
        return "checkbox"
    if raw in {"spreadsheet", "table"}:
        return "spreadsheet"
    # Fallback keeps unexpected values visible for debugging instead of hiding data.
    return raw or "text"


def to_text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text if text else None


def build_guide(row: dict[str, Any]) -> str | None:
    evidence_source = to_text(row.get("Evidence Source"))
    collection_method = to_text(row.get("Collection Method"))
    aws_auto_level = to_text(row.get("AWS Auto Level"))
    aws_services = to_text(row.get("AWS Services"))
    aws_question_sources = to_text(row.get("Question-Level AWS Sources"))
    rationale = to_text(row.get("Reason / Rationale"))

    parts: list[str] = []
    if evidence_source:
        parts.append(f"Evidence source: {evidence_source}")
    if collection_method:
        parts.append(f"Collection method: {collection_method}")
    if aws_auto_level:
        parts.append(f"AWS auto level: {aws_auto_level}")
    if aws_services:
        parts.append(f"AWS services: {aws_services}")
    if aws_question_sources:
        parts.append(f"Question-level AWS sources: {aws_question_sources}")
    if rationale:
        parts.append(f"Reason / rationale: {rationale}")
    return "\n".join(parts) if parts else None


def read_source_rows(file_path: Path) -> list[dict[str, Any]]:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return []

        header_idx = None
        headers: list[str] = []
        for i, row in enumerate(all_rows):
            candidate = [str(c or "").strip() for c in row]
            candidate_norm = {normalize_header(c) for c in candidate if c}
            expected_norm = {normalize_header(h) for h in EXPECTED_HEADERS}
            if expected_norm.issubset(candidate_norm):
                header_idx = i
                headers = candidate
                break

        if header_idx is None:
            return []

        data_rows = all_rows[header_idx + 1 :]
        return [dict(zip(headers, row)) for row in data_rows if any(cell is not None and str(cell).strip() for cell in row)]

    if suffix == ".csv":
        with file_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return [dict(r) for r in reader if any(str(v or "").strip() for v in r.values())]

    raise ValueError(f"Unsupported file type: {file_path.suffix}. Use .xlsx or .csv.")


def normalize_row_keys(row: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in row.items():
        source_key = normalize_header(str(key))
        canonical = HEADER_MAP.get(source_key)
        if canonical:
            out[canonical] = value
    return out


def validate_headers(rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError("Source file has no data rows.")
    sample = rows[0]
    missing = [h for h in EXPECTED_HEADERS if h not in sample]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")


def db_connect():
    load_dotenv(BACKEND_DIR / ".env")
    return psycopg2.connect(
        host=os.getenv("DB_HOST", "127.0.0.1"),
        port=int(os.getenv("DB_PORT", "5432")),
        dbname=os.getenv("DB_NAME", "compliance"),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASSWORD", ""),
        sslmode="require" if os.getenv("DB_SSL", "false").lower() == "true" else "disable",
    )


def fetch_valid_evidence_items(cur) -> set[str]:
    cur.execute("SELECT id FROM swift_2026.canonical_evidence_items")
    return {str(r[0]).strip() for r in cur.fetchall()}


def build_payload(rows: list[dict[str, Any]], valid_evidence_items: set[str]) -> tuple[list[tuple], list[str]]:
    sort_order_by_item: defaultdict[str, int] = defaultdict(int)
    payload: list[tuple] = []
    skipped: list[str] = []

    for idx, row in enumerate(rows, start=2):
        control_id = to_text(row.get("Control ID"))
        evidence_item_raw = to_text(row.get("Evidence Item"))
        question_key = to_text(row.get("Question Key"))
        label = to_text(row.get("Question (Label)"))
        question_type = normalize_question_type(row.get("Question Type"))
        required = parse_required(row.get("Evidence Required"), question_type)

        if not evidence_item_raw or not question_key or not label:
            skipped.append(f"Row {idx}: missing Evidence Item / Question Key / Question (Label)")
            continue
        item_ids = [x.strip() for x in evidence_item_raw.replace("+", ",").split(",") if x and x.strip()]
        if not item_ids:
            skipped.append(f"Row {idx}: invalid Evidence Item value '{evidence_item_raw}'")
            continue

        rows_value = 3 if question_type == "textarea" else None
        guide = build_guide(row)

        for evidence_item_id in item_ids:
            if evidence_item_id not in valid_evidence_items:
                skipped.append(f"Row {idx}: unknown evidence_item_id '{evidence_item_id}' (not in canonical_evidence_items)")
                continue

            sort_order = sort_order_by_item[evidence_item_id]
            sort_order_by_item[evidence_item_id] += 1

            payload.append(
                (
                    evidence_item_id,
                    question_key,
                    label,
                    question_type,
                    required,
                    None,  # placeholder
                    json.dumps([]),  # options
                    sort_order,
                    control_id,
                    rows_value,
                    None,  # accept
                    None,  # upload_label
                    "2026v",
                    guide,
                    None,  # show_when_question
                    json.dumps([]),  # show_when_values
                )
            )
    return payload, skipped


def run_import(file_path: Path, truncate_first: bool = False) -> int:
    raw_rows = read_source_rows(file_path)
    rows = [normalize_row_keys(r) for r in raw_rows]
    validate_headers(rows)

    conn = db_connect()
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            valid_items = fetch_valid_evidence_items(cur)
            payload, skipped = build_payload(rows, valid_items)

            if truncate_first:
                cur.execute("TRUNCATE TABLE swift_2026.evidence_based_questions RESTART IDENTITY CASCADE")

            upsert_sql = """
                INSERT INTO swift_2026.evidence_based_questions (
                    evidence_item_id, question_key, label, question_type, required, placeholder,
                    options, sort_order, control_id, rows, accept, upload_label, cscf_version,
                    guide, show_when_question, show_when_values
                ) VALUES (
                    %s, %s, %s, %s, %s, %s,
                    %s::jsonb, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s::jsonb
                )
                ON CONFLICT (evidence_item_id, question_key) DO UPDATE SET
                    label = EXCLUDED.label,
                    question_type = EXCLUDED.question_type,
                    required = EXCLUDED.required,
                    placeholder = EXCLUDED.placeholder,
                    options = EXCLUDED.options,
                    sort_order = EXCLUDED.sort_order,
                    control_id = EXCLUDED.control_id,
                    rows = EXCLUDED.rows,
                    accept = EXCLUDED.accept,
                    upload_label = EXCLUDED.upload_label,
                    cscf_version = EXCLUDED.cscf_version,
                    guide = EXCLUDED.guide,
                    show_when_question = EXCLUDED.show_when_question,
                    show_when_values = EXCLUDED.show_when_values
            """
            execute_batch(cur, upsert_sql, payload, page_size=500)
            conn.commit()

            print(f"[OK] Source rows read: {len(rows)}")
            print(f"[OK] Upserted rows:    {len(payload)}")
            print(f"[OK] Skipped rows:     {len(skipped)}")
            if skipped:
                print("\nSkipped details:")
                for msg in skipped[:100]:
                    print(f"  - {msg}")
                if len(skipped) > 100:
                    print(f"  ... and {len(skipped) - 100} more")
        return 0
    except Exception as exc:
        conn.rollback()
        print(f"[ERROR] Import failed: {exc}", file=sys.stderr)
        return 1
    finally:
        conn.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Import SWIFT CSCF 2026 evidence questions into PostgreSQL.")
    parser.add_argument(
        "--file",
        type=str,
        default=str(DEFAULT_INPUT_FILE),
        help="Path to source .xlsx or .csv file.",
    )
    parser.add_argument(
        "--truncate-first",
        action="store_true",
        help="Truncate swift_2026.evidence_based_questions before upsert.",
    )
    args = parser.parse_args()

    source_file = Path(args.file).resolve()
    if not source_file.exists():
        print(f"[ERROR] File not found: {source_file}", file=sys.stderr)
        return 1

    print(f"[INFO] Importing from: {source_file}")
    return run_import(source_file, truncate_first=args.truncate_first)


if __name__ == "__main__":
    sys.exit(main())
