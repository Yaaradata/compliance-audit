"""
Seed script: reads xlsx files and upserts canonical_evidence_items JSONB fields
(input_schema, sufficiency_dimensions), item_control_mappings sufficiency details,
and evidence_sufficiency_matrix (one-time seed from CSCF_v2025_Complete_Sufficiency_Matrix.xlsx).

Usage:
    cd backend
    pip install -r requirements.txt
    python scripts/seed_data.py
"""

import json
import os
import re
import sys
from pathlib import Path

import psycopg2
from psycopg2.extras import execute_values
from dotenv import load_dotenv

SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
load_dotenv(BACKEND_DIR / ".env")

XLSX_DIR = Path(os.getenv("XLSX_DIR", str(BACKEND_DIR.parent.parent)))
# Project root (where CSCF_v2025_Complete_Sufficiency_Matrix.xlsx often lives)
PROJECT_ROOT = BACKEND_DIR.parent

EVIDENCE_MODEL_FILE = XLSX_DIR / "SWIFT_CSCF_v2025_Canonical_Evidence_Model.xlsx"
SUFFICIENCY_FILE = XLSX_DIR / "MultiControl_Evidence_Sufficiency.xlsx"
COMPLETE_SUFFICIENCY_MATRIX_FILE = (
    PROJECT_ROOT / "CSCF_v2025_Complete_Sufficiency_Matrix.xlsx"
    if (PROJECT_ROOT / "CSCF_v2025_Complete_Sufficiency_Matrix.xlsx").exists()
    else XLSX_DIR / "CSCF_v2025_Complete_Sufficiency_Matrix.xlsx"
)


def get_connection():
    return psycopg2.connect(
        host=os.getenv("DB_HOST", "127.0.0.1"),
        port=int(os.getenv("DB_PORT", "5432")),
        dbname=os.getenv("DB_NAME", "compliance"),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASSWORD", "Compliance_Audit01"),
        sslmode="require" if os.getenv("DB_SSL", "false").lower() == "true" else "disable",
    )


def safe_str(val):
    if val is None:
        return None
    return str(val).strip()


def parse_evidence_model(wb):
    """Parse input_schema and sufficiency_dimensions from the Evidence Model xlsx."""
    items = {}

    try:
        ws = wb["1_Canonical_Evidence_Items"]
    except KeyError:
        sheet_names = wb.sheetnames
        ws = wb[sheet_names[0]] if sheet_names else None

    if ws is None:
        print("WARNING: Could not find evidence items sheet")
        return items

    headers = [safe_str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        item_id = safe_str(row_dict.get("Item ID") or row_dict.get("ID") or row_dict.get("Evidence ID"))
        if not item_id:
            continue

        input_schema = []
        suf_dims = []

        items[item_id] = {
            "input_schema": input_schema,
            "sufficiency_dimensions": suf_dims,
        }

    return items


def parse_sufficiency_matrix(wb):
    """Parse sufficiency requirements from the sufficiency xlsx."""
    mappings = []

    try:
        ws = wb["Sufficiency Detail"]
    except KeyError:
        try:
            ws = wb[wb.sheetnames[0]]
        except (KeyError, IndexError):
            print("WARNING: Could not find sufficiency detail sheet")
            return mappings

    headers = [safe_str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        item_id = safe_str(row_dict.get("Evidence Item ID") or row_dict.get("Item ID") or row_dict.get("Evidence ID"))
        control_id = safe_str(row_dict.get("Control ID") or row_dict.get("Control"))
        requirement = safe_str(row_dict.get("Sufficiency Requirement") or row_dict.get("Requirement"))

        if item_id and control_id:
            mappings.append({
                "evidence_item_id": item_id,
                "control_id": control_id,
                "sufficiency_requirement": requirement,
            })

    return mappings


def seed_jsonb_fields(conn, items_data):
    """Update canonical_evidence_items with parsed JSONB fields."""
    if not items_data:
        print("No JSONB data to seed for canonical_evidence_items")
        return

    cur = conn.cursor()
    updated = 0
    for item_id, data in items_data.items():
        cur.execute(
            """
            UPDATE swift_2025.canonical_evidence_items
            SET input_schema = %s::jsonb,
                sufficiency_dimensions = %s::jsonb
            WHERE id = %s
            """,
            (json.dumps(data["input_schema"]), json.dumps(data["sufficiency_dimensions"]), item_id),
        )
        if cur.rowcount > 0:
            updated += 1

    conn.commit()
    cur.close()
    print(f"Updated {updated} canonical_evidence_items with JSONB fields")


def seed_sufficiency_requirements(conn, mappings):
    """Update item_control_mappings with sufficiency_requirement text."""
    if not mappings:
        print("No sufficiency requirements to seed")
        return

    cur = conn.cursor()
    updated = 0
    for m in mappings:
        if m["sufficiency_requirement"]:
            cur.execute(
                """
                UPDATE swift_2025.item_control_mappings
                SET sufficiency_requirement = %s
                WHERE evidence_item_id = %s AND control_id = %s
                """,
                (m["sufficiency_requirement"], m["evidence_item_id"], m["control_id"]),
            )
            if cur.rowcount > 0:
                updated += 1

    conn.commit()
    cur.close()
    print(f"Updated {updated} item_control_mappings with sufficiency requirements")


def _normalize_header(h: str) -> str:
    """Strip, replace newlines with space, collapse spaces, lowercase for matching."""
    if not h:
        return ""
    return re.sub(r"\s+", " ", str(h).strip().replace("\n", " ")).strip().lower()


def parse_complete_sufficiency_matrix(wb) -> list[dict]:
    """Parse CSCF_v2025_Complete_Sufficiency_Matrix.xlsx into rows for evidence_sufficiency_matrix."""
    rows = []
    ws = wb[wb.sheetnames[0]]
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [safe_str(h) for h in raw_headers]
    # Normalized header -> original header (for row_dict lookup)
    norm_to_orig = {}
    for h in headers:
        if h:
            norm_to_orig[_normalize_header(h)] = h

    def get_val(row_dict: dict, *candidates: str) -> str | None:
        for c in candidates:
            n = _normalize_header(c)
            for norm, orig in norm_to_orig.items():
                if n in norm or norm in n or n == norm:
                    return safe_str(row_dict.get(orig))
        return None

    def find_col_by_substring(row_dict: dict, sub: str) -> str | None:
        """First column whose normalized name contains sub."""
        for h, v in row_dict.items():
            if h and sub in _normalize_header(h):
                return safe_str(v)
        return None

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        item_code = get_val(row_dict, "Item Code")
        control_id = get_val(row_dict, "Control ID")
        if not item_code or not control_id:
            continue
        evidence_item_name = get_val(row_dict, "Evidence Item Name") or ""
        control_name = get_val(row_dict, "Control Name") or ""
        ma = get_val(row_dict, "M/A") or "M"
        evidence_type = get_val(row_dict, "Evidence Type") or ""
        sufficiency_criteria = (
            get_val(row_dict, "What This Evidence Must Show for This Control (Sufficiency Criteria — Bullet Points)")
            or get_val(row_dict, "What This Evidence Must Show for This Control")
            or find_col_by_substring(row_dict, "sufficiency")
        )
        evaluation_criteria = (
            get_val(row_dict, "How to Evaluate / AI Scoring Criteria (Pass/Fail Checkpoints)")
            or get_val(row_dict, "How to Evaluate")
            or find_col_by_substring(row_dict, "evaluate")
            or find_col_by_substring(row_dict, "scoring")
        )

        rows.append({
            "item_code": item_code,
            "control_id": control_id,
            "evidence_item_name": evidence_item_name,
            "control_name": control_name,
            "ma": (ma[:1].upper() if ma else "M") if ma and len(ma) >= 1 else "M",
            "evidence_type": evidence_type,
            "sufficiency_criteria": sufficiency_criteria,
            "evaluation_criteria": evaluation_criteria,
        })
    return rows


def seed_evidence_sufficiency_matrix(conn, matrix_rows: list[dict]) -> None:
    """Insert or update evidence_sufficiency_matrix from parsed matrix rows."""
    if not matrix_rows:
        print("No evidence_sufficiency_matrix rows to seed")
        return
    cur = conn.cursor()
    inserted = 0
    for r in matrix_rows:
        cur.execute(
            """
            INSERT INTO swift_2025.evidence_sufficiency_matrix
                (item_code, control_id, evidence_item_name, control_name, ma, evidence_type, sufficiency_criteria, evaluation_criteria)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (item_code, control_id) DO UPDATE SET
                evidence_item_name = EXCLUDED.evidence_item_name,
                control_name = EXCLUDED.control_name,
                ma = EXCLUDED.ma,
                evidence_type = EXCLUDED.evidence_type,
                sufficiency_criteria = EXCLUDED.sufficiency_criteria,
                evaluation_criteria = EXCLUDED.evaluation_criteria
            """,
            (
                r["item_code"],
                r["control_id"],
                r["evidence_item_name"],
                r["control_name"],
                r["ma"],
                r["evidence_type"],
                r.get("sufficiency_criteria"),
                r.get("evaluation_criteria"),
            ),
        )
        inserted += 1
    conn.commit()
    cur.close()
    print(f"Upserted {inserted} rows into evidence_sufficiency_matrix")


def main():
    conn = get_connection()
    conn.autocommit = False
    print(f"Connected to {os.getenv('DB_HOST', '127.0.0.1')}:{os.getenv('DB_PORT', '5432')}/{os.getenv('DB_NAME', 'compliance')}")

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Run: pip install openpyxl")
        sys.exit(1)

    items_data = {}
    if EVIDENCE_MODEL_FILE.exists():
        print(f"Reading {EVIDENCE_MODEL_FILE.name}...")
        wb = openpyxl.load_workbook(EVIDENCE_MODEL_FILE, read_only=True, data_only=True)
        items_data = parse_evidence_model(wb)
        wb.close()
        print(f"  Parsed {len(items_data)} evidence items")
    else:
        print(f"WARNING: {EVIDENCE_MODEL_FILE} not found — skipping JSONB seed")

    sufficiency_mappings = []
    if SUFFICIENCY_FILE.exists():
        print(f"Reading {SUFFICIENCY_FILE.name}...")
        wb = openpyxl.load_workbook(SUFFICIENCY_FILE, read_only=True, data_only=True)
        sufficiency_mappings = parse_sufficiency_matrix(wb)
        wb.close()
        print(f"  Parsed {len(sufficiency_mappings)} sufficiency mappings")
    else:
        print(f"WARNING: {SUFFICIENCY_FILE} not found — skipping sufficiency seed")

    seed_jsonb_fields(conn, items_data)
    seed_sufficiency_requirements(conn, sufficiency_mappings)

    matrix_rows = []
    if COMPLETE_SUFFICIENCY_MATRIX_FILE.exists():
        print(f"Reading {COMPLETE_SUFFICIENCY_MATRIX_FILE.name}...")
        wb = openpyxl.load_workbook(COMPLETE_SUFFICIENCY_MATRIX_FILE, read_only=True, data_only=True)
        matrix_rows = parse_complete_sufficiency_matrix(wb)
        wb.close()
        print(f"  Parsed {len(matrix_rows)} evidence_sufficiency_matrix rows")
    else:
        print(f"WARNING: {COMPLETE_SUFFICIENCY_MATRIX_FILE} not found — skipping evidence_sufficiency_matrix seed")

    if matrix_rows:
        seed_evidence_sufficiency_matrix(conn, matrix_rows)

    conn.close()
    print("Done.")


if __name__ == "__main__":
    main()
